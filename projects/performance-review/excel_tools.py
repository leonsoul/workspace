#!/usr/bin/env python3
"""
Excel 导入导出工具

功能:
- 导出人员数据到 Excel
- 导出评价数据到 Excel
- 从 Excel 导入人员数据
- 从 Excel 导入评价数据
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from pathlib import Path
import io

# 导入数据库模型
import sys
sys.path.insert(0, '/home/admin/.openclaw/workspace/projects/performance-review')
from database import Member, Review, Issue, SessionLocal


def get_session():
    return SessionLocal()


# ============== 样式定义 ==============

def create_header_style():
    """创建表头样式"""
    return {
        'font': Font(bold=True, color='FFFFFF', size=12),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    }

def create_cell_style():
    """创建单元格样式"""
    return {
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
    }


# ============== 导出功能 ==============

def export_members_to_excel():
    """导出人员数据到 Excel"""
    db = get_session()
    try:
        members = db.query(Member).filter_by(is_active=True).all()
        
        # 准备数据
        data = []
        for m in members:
            data.append({
                'ID': m.id,
                '姓名': m.name,
                '邮箱': m.email,
                '部门': m.department,
                '评价次数': len(m.reviews),
                '平均分': m.get_average_score(),
                '创建时间': m.created_at.strftime('%Y-%m-%d %H:%M:%S') if m.created_at else ''
            })
        
        # 创建 DataFrame
        df = pd.DataFrame(data)
        
        # 保存到字节流
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='人员列表', index=False)
            
            # 获取 worksheet
            ws = writer.sheets['人员列表']
            
            # 设置列宽
            column_widths = [8, 15, 25, 15, 12, 10, 20]
            for i, width in enumerate(column_widths, 1):
                col_letter = chr(64 + i)  # A, B, C...
                ws.column_dimensions[col_letter].width = width
            
            # 设置表头样式
            header_style = create_header_style()
            for cell in ws[1]:
                for key, value in header_style.items():
                    setattr(cell, key, value)
        
        output.seek(0)
        return output.getvalue()
    
    finally:
        db.close()


def export_reviews_to_excel(member_id=None, period=None):
    """导出评价数据到 Excel"""
    db = get_session()
    try:
        query = db.query(Review)
        if member_id:
            query = query.filter_by(member_id=member_id)
        if period:
            query = query.filter_by(period=period)
        
        reviews = query.order_by(Review.created_at.desc()).all()
        
        # 准备数据
        data = []
        for r in reviews:
            data.append({
                'ID': r.id,
                '人员姓名': r.member.name if r.member else '',
                '周期': r.period,
                '分数': r.score,
                '等级': r.level,
                '评价': r.description,
                '问题数': len(r.issues),
                '评价人': r.created_by,
                '创建时间': r.created_at.strftime('%Y-%m-%d %H:%M:%S') if r.created_at else ''
            })
        
        df = pd.DataFrame(data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='评价记录', index=False)
            
            ws = writer.sheets['评价记录']
            
            # 设置列宽
            column_widths = [8, 15, 12, 8, 12, 30, 8, 15, 20]
            for i, width in enumerate(column_widths, 1):
                col_letter = chr(64 + i)
                ws.column_dimensions[col_letter].width = width
            
            # 设置表头样式
            header_style = create_header_style()
            for cell in ws[1]:
                for key, value in header_style.items():
                    setattr(cell, key, value)
        
        output.seek(0)
        return output.getvalue()
    
    finally:
        db.close()


# ============== 导入功能 ==============

def import_members_from_excel(file_content):
    """从 Excel 导入人员数据"""
    db = get_session()
    try:
        # 读取 Excel
        df = pd.read_excel(io.BytesIO(file_content), sheet_name=0)
        
        results = {
            'success': 0,
            'failed': 0,
            'errors': []
        }
        
        for index, row in df.iterrows():
            try:
                # 跳过表头
                if index == 0 and '姓名' not in str(row.get('姓名', '')):
                    continue
                
                name = str(row.get('姓名', '')).strip()
                if not name:
                    results['failed'] += 1
                    results['errors'].append(f'行{index+1}: 姓名为空')
                    continue
                
                # 检查重名
                existing = db.query(Member).filter_by(name=name).first()
                if existing:
                    results['failed'] += 1
                    results['errors'].append(f'行{index+1}: 人员"{name}"已存在')
                    continue
                
                # 创建人员
                member = Member(
                    name=name,
                    email=str(row.get('邮箱', '')).strip(),
                    department=str(row.get('部门', '')).strip()
                )
                db.add(member)
                results['success'] += 1
                
            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f'行{index+1}: {str(e)}')
        
        db.commit()
        return results
    
    except Exception as e:
        db.rollback()
        raise Exception(f'导入失败：{str(e)}')
    finally:
        db.close()


def import_reviews_from_excel(file_content):
    """从 Excel 导入评价数据"""
    from database import calculate_score as calc_score
    
    db = get_session()
    try:
        df = pd.read_excel(io.BytesIO(file_content), sheet_name=0)
        
        results = {
            'success': 0,
            'failed': 0,
            'errors': []
        }
        
        for index, row in df.iterrows():
            try:
                member_name = str(row.get('人员姓名', '')).strip()
                if not member_name:
                    results['failed'] += 1
                    results['errors'].append(f'行{index+1}: 人员姓名为空')
                    continue
                
                # 查找人员
                member = db.query(Member).filter_by(name=member_name).first()
                if not member:
                    results['failed'] += 1
                    results['errors'].append(f'行{index+1}: 人员"{member_name}"不存在')
                    continue
                
                # 计算评分
                score = float(row.get('分数', 5.0))
                level = str(row.get('等级', '正常')).strip()
                
                # 创建评价
                review = Review(
                    member_id=member.id,
                    period=str(row.get('周期', datetime.now().strftime('%Y-%m'))).strip(),
                    score=score,
                    level=level,
                    description=str(row.get('评价', '')).strip(),
                    created_by=str(row.get('评价人', '导入')).strip()
                )
                db.add(review)
                db.commit()
                
                results['success'] += 1
                
            except Exception as e:
                db.rollback()
                results['failed'] += 1
                results['errors'].append(f'行{index+1}: {str(e)}')
        
        return results
    
    except Exception as e:
        db.rollback()
        raise Exception(f'导入失败：{str(e)}')
    finally:
        db.close()


# ============== 模板生成 ==============

def generate_members_template():
    """生成人员导入模板"""
    data = [{
        '姓名': '张三',
        '邮箱': 'zhangsan@example.com',
        '部门': '研发部'
    }]
    
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='人员模板', index=False)
        
        ws = writer.sheets['人员模板']
        
        # 添加说明
        ws['A4'] = '说明：'
        ws['A5'] = '1. 姓名必填，不能重复'
        ws['A6'] = '2. 邮箱和部门选填'
        ws['A7'] = '3. 不要删除表头行'
    
    output.seek(0)
    return output.getvalue()


def generate_reviews_template():
    """生成评价导入模板"""
    data = [{
        '人员姓名': '张三',
        '周期': '2026-03',
        '分数': 4.5,
        '等级': '优秀',
        '评价': '工作表现优秀，代码质量高',
        '评价人': '李四'
    }]
    
    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='评价模板', index=False)
        
        ws = writer.sheets['评价模板']
        
        # 添加说明
        ws['A7'] = '说明：'
        ws['A8'] = '1. 人员姓名必填，必须是已存在的人员'
        ws['A9'] = '2. 周期格式：YYYY-MM (如：2026-03)'
        ws['A10'] = '3. 分数范围：0-5'
        ws['A11'] = '4. 等级：卓越/优秀/良好/正常/及格/质量差'
    
    output.seek(0)
    return output.getvalue()


if __name__ == '__main__':
    # 测试导出
    print("测试导出人员...")
    data = export_members_to_excel()
    print(f"✅ 导出成功，{len(data)} 字节")
    
    print("\n测试导出评价...")
    data = export_reviews_to_excel()
    print(f"✅ 导出成功，{len(data)} 字节")
